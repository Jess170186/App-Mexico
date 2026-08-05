#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
App Mexico — Extraction mensuelle GRATUITE via OpenStreetMap
===============================================================
Aucune clé, aucune carte bancaire, aucune limite. Une fois par mois, ce script
interroge OpenStreetMap (API Overpass) pour TOUT le Quintana Roo et le Yucatán,
et extrait restaurants, bars, commerces, plages, activités et transports avec :
    nom · adresse · description · contact (tél / site) · catégorie · ville · coordonnées

Il écrit DEUX fichiers :
    • repertoire.xlsx  (tableur lisible/éditable : la « base » que tu voulais)
    • places.json      (le format que lit l'app App Mexico)

Robustesse : on interroge par « bounding box » (pas par nom de zone, plus fiable),
on RÉESSAYE automatiquement sur plusieurs serveurs Overpass miroirs, et on garde
TOUT lieu qui a un nom (le filtre « qualité » ne rejette plus les lieux sans site web).

Utilisation :
    pip install -r requirements.txt
    python osm_extract.py
"""

import json, time, math, pathlib
import requests
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

HERE = pathlib.Path(__file__).parent

# Plusieurs serveurs Overpass : on essaie le suivant si l'un est occupé / en panne.
ENDPOINTS = [
    "https://overpass-api.de/api/interpreter",
    "https://overpass.kumi.systems/api/interpreter",
    "https://overpass.private.coffee/api/interpreter",
    "https://maps.mail.ru/osm/tools/overpass/api/interpreter",
]

# Bounding box couvrant tout le Quintana Roo + Yucatán (+ marge).
# (sud, ouest, nord, est)
BBOX = (17.70, -90.80, 21.75, -86.55)

# Villes de référence (pour rattacher chaque lieu à une ville et à une région)
CITY = {
    "Cancún": ("Quintana Roo", 21.161, -86.851), "Playa del Carmen": ("Quintana Roo", 20.629, -87.073),
    "Tulum": ("Quintana Roo", 20.211, -87.465), "Bacalar": ("Quintana Roo", 18.680, -88.395),
    "Cozumel": ("Quintana Roo", 20.422, -86.922), "Holbox": ("Quintana Roo", 21.522, -87.379),
    "Chetumal": ("Quintana Roo", 18.503, -88.305), "Puerto Morelos": ("Quintana Roo", 20.848, -86.875),
    "Isla Mujeres": ("Quintana Roo", 21.232, -86.731), "Akumal": ("Quintana Roo", 20.393, -87.315),
    "Puerto Aventuras": ("Quintana Roo", 20.499, -87.226), "Mahahual": ("Quintana Roo", 18.716, -87.712),
    "Felipe Carrillo Puerto": ("Quintana Roo", 19.578, -88.045), "Kantunilkín": ("Quintana Roo", 21.100, -87.490),
    "Mérida": ("Yucatán", 20.967, -89.624), "Valladolid": ("Yucatán", 20.688, -88.202),
    "Progreso": ("Yucatán", 21.282, -89.665), "Izamal": ("Yucatán", 20.934, -89.017),
    "Tizimín": ("Yucatán", 21.143, -88.152), "Uxmal": ("Yucatán", 20.360, -89.771),
    "Ticul": ("Yucatán", 20.398, -89.534), "Tekax": ("Yucatán", 20.201, -89.286),
    "Motul": ("Yucatán", 21.095, -89.285), "Chichén Itzá": ("Yucatán", 20.684, -88.568),
    "Celestún": ("Yucatán", 20.858, -90.383),
}
MAX_KM = 60  # un lieu est gardé s'il est à moins de 60 km d'une ville connue

# Catégorie app -> filtres OSM (node + way pour capter plages, sites, etc.)
CATS = {
    "Restaurantes": ['node["amenity"="restaurant"]', 'node["amenity"="cafe"]',
                     'node["amenity"="fast_food"]'],
    "Bares":        ['node["amenity"="bar"]', 'node["amenity"="pub"]', 'node["amenity"="nightclub"]',
                     'node["amenity"="biergarten"]'],
    "Compras":      ['node["shop"="mall"]', 'node["shop"="department_store"]',
                     'node["amenity"="marketplace"]', 'node["shop"="gift"]',
                     'node["shop"="supermarket"]', 'node["shop"="convenience"]',
                     'node["shop"="clothes"]', 'node["shop"="shoes"]',
                     'node["shop"="bakery"]', 'node["shop"="jewelry"]',
                     'node["shop"="chemist"]', 'node["amenity"="pharmacy"]',
                     'node["shop"="greengrocer"]', 'node["shop"="butcher"]',
                     'way["shop"="mall"]', 'way["shop"="department_store"]'],
    "Playas":       ['node["natural"="beach"]', 'way["natural"="beach"]'],
    "Actividades":  ['node["tourism"="attraction"]', 'node["tourism"="museum"]',
                     'node["tourism"="theme_park"]', 'way["tourism"="attraction"]',
                     'way["historic"="archaeological_site"]', 'node["historic"="archaeological_site"]',
                     'node["natural"="cenote"]', 'node["water"="cenote"]'],
    "Transporte":   ['node["amenity"="bus_station"]', 'node["amenity"="ferry_terminal"]',
                     'node["aeroway"="aerodrome"]', 'way["aeroway"="aerodrome"]'],
}

def km(a1, a2, b1, b2):
    R = 6371; dLat = math.radians(b1 - a1); dLng = math.radians(b2 - a2)
    s = math.sin(dLat/2)**2 + math.cos(math.radians(a1))*math.cos(math.radians(b1))*math.sin(dLng/2)**2
    return 2*R*math.asin(math.sqrt(s))

def nearest_city(lat, lng):
    best = None
    for c, (r, la, lo) in CITY.items():
        d = km(lat, lng, la, lo)
        if not best or d < best[2]:
            best = (c, r, d)
    return best  # (city, region, dist_km)

def overpass(filters):
    """Interroge Overpass par bounding box, avec réessais sur plusieurs miroirs."""
    s, w, n, e = BBOX
    body = "[out:json][timeout:180];\n(\n"
    for f in filters:
        body += f"  {f}({s},{w},{n},{e});\n"
    body += ");\nout center tags;"
    last_err = None
    for attempt in range(3):
        for url in ENDPOINTS:
            try:
                r = requests.post(url, data={"data": body},
                                  headers={"User-Agent": "AppMexico/1.0 (contact: marketing@saleahomes.com)"},
                                  timeout=200)
                if r.status_code in (429, 502, 503, 504):
                    last_err = f"{url} -> HTTP {r.status_code}"
                    time.sleep(5); continue
                r.raise_for_status()
                return r.json().get("elements", [])
            except Exception as ex:
                last_err = f"{url} -> {ex}"
                time.sleep(3); continue
        time.sleep(8)  # tous les miroirs occupés : on patiente avant un nouveau tour
    raise RuntimeError(f"Overpass indisponible après plusieurs essais ({last_err})")

def describe(cat, tags):
    if cat == "Restaurantes":
        cu = tags.get("cuisine", "").replace("_", " ")
        return ("Restaurante · " + cu) if cu else "Restaurante"
    if cat == "Bares":     return "Bar / vida nocturna"
    if cat == "Compras":   return tags.get("shop", tags.get("amenity", "Comercio")).replace("_", " ").title()
    if cat == "Playas":    return "Playa"
    if cat == "Actividades":
        return (tags.get("tourism") or tags.get("historic") or tags.get("natural")
                or "Atracción").replace("_", " ").title()
    if cat == "Transporte":
        return {"bus_station": "Terminal de autobuses", "ferry_terminal": "Terminal de ferry"}.get(
            tags.get("amenity"), "Aeropuerto" if tags.get("aeroway") else "Transporte")
    return cat

def address(tags):
    parts = [tags.get("addr:street", ""), tags.get("addr:housenumber", ""),
             tags.get("addr:city", "")]
    return " ".join(p for p in parts if p).strip()

def contact(tags):
    return (tags.get("phone") or tags.get("contact:phone") or ""), \
           (tags.get("website") or tags.get("contact:website") or "")

def social(tags):
    def norm(v, base):
        if not v:
            return ""
        return v if v.startswith("http") else base + v.lstrip("@/")
    fb = norm(tags.get("contact:facebook") or tags.get("facebook") or "", "https://facebook.com/")
    ig = norm(tags.get("contact:instagram") or tags.get("instagram") or "", "https://instagram.com/")
    return fb, ig

def gmaps_url(lat, lng):
    return f"https://www.google.com/maps/search/?api=1&query={lat},{lng}"

def run():
    records, seen = [], set()
    for cat, filters in CATS.items():
        print(f"→ {cat} …", flush=True)
        try:
            els = overpass(filters)
        except Exception as ex:
            print(f"   ⚠️ {cat} : {ex}", flush=True); continue
        added = 0
        for el in els:
            tags = el.get("tags", {})
            name = tags.get("name")
            if not name:
                continue
            lat = el.get("lat") or (el.get("center") or {}).get("lat")
            lng = el.get("lon") or (el.get("center") or {}).get("lon")
            if lat is None or lng is None:
                continue
            city, region, dist = nearest_city(lat, lng)
            if dist > MAX_KM:
                continue
            key = (name.lower(), round(lat, 4), round(lng, 4))
            if key in seen:
                continue
            seen.add(key)
            phone, website = contact(tags)
            fb, ig = social(tags)
            records.append({
                "n": name, "c": cat, "city": city, "region": region,
                "d": describe(cat, tags), "address": address(tags), "phone": phone,
                "website": website, "facebook": fb, "instagram": ig,
                "gmaps": gmaps_url(round(lat, 6), round(lng, 6)),
                "img": "", "w": tags.get("internet_access") in ("wlan", "yes"),
                "sub": (tags.get("shop") or tags.get("amenity") or tags.get("tourism") or ""),
                "lat": round(lat, 6), "lng": round(lng, 6),
            })
            added += 1
        print(f"   {added} lieux gardés", flush=True)
        time.sleep(4)  # courtoisie envers les serveurs Overpass

    records.sort(key=lambda r: (r["region"], r["city"], r["c"], r["n"]))

    # ---- places.json (pour l'app) ----
    (HERE / "places.json").write_text(json.dumps(
        {"updated_at": time.strftime("%Y-%m-%dT%H:%M:%SZ", time.gmtime()),
         "count": len(records), "source": "OpenStreetMap", "places": records},
        ensure_ascii=False, indent=2), encoding="utf-8")

    # ---- repertoire.xlsx (le tableur) ----
    wb = Workbook(); ws = wb.active; ws.title = "Répertoire"
    head = ["Nom", "Catégorie", "Ville", "Région", "Adresse", "Description",
            "Téléphone", "Site web", "Facebook", "Instagram", "Google Maps",
            "Image URL", "Note", "Lat", "Lng"]
    ws.append(head)
    for i, h in enumerate(head, 1):
        c = ws.cell(row=1, column=i); c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="0F7D7B"); c.alignment = Alignment(vertical="center")
    for r in records:
        ws.append([r["n"], r["c"], r["city"], r["region"], r["address"], r["d"],
                   r["phone"], r["website"], r["facebook"], r["instagram"], r["gmaps"],
                   r["img"], "—", r["lat"], r["lng"]])
    widths = [26, 14, 16, 14, 30, 22, 16, 26, 26, 26, 34, 16, 6, 10, 10]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64+i)].width = w
    ws.freeze_panes = "A2"
    wb.save(HERE / "repertoire.xlsx")

    print(f"\n✅ {len(records)} lieux → repertoire.xlsx + places.json", flush=True)

if __name__ == "__main__":
    run()
